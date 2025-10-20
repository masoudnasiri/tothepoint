"""
Reports & Analytics Router
Provides comprehensive data aggregation and analytics for the Reports & Analytics page.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_, or_, extract
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.auth import get_current_user
from app.models import User, FinalizedDecision, CashflowEvent, Project, Supplier
from app.schemas import UserRead

router = APIRouter(prefix="/reports", tags=["Reports & Analytics"])


def calculate_percentile(values: List[float], percentile: int) -> float:
    """Calculate percentile from a list of values."""
    if not values:
        return 0.0
    sorted_values = sorted(values)
    index = (percentile / 100) * (len(sorted_values) - 1)
    lower = int(index)
    upper = lower + 1
    if upper >= len(sorted_values):
        return sorted_values[-1]
    weight = index - lower
    return sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight


def aggregate_financial_summary(
    db: Session,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_ids: Optional[List[int]]
) -> Dict[str, Any]:
    """Aggregate financial summary data: cash flow and budget vs actuals."""
    
    # Base query for finalized decisions
    decisions_query = db.query(FinalizedDecision).filter(
        FinalizedDecision.status == 'FINALIZED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.filter(
            or_(
                FinalizedDecision.decision_finalized_at >= datetime.combine(start_date, datetime.min.time()),
                FinalizedDecision.planned_purchase_date >= start_date
            )
        )
    if end_date:
        decisions_query = decisions_query.filter(
            or_(
                FinalizedDecision.decision_finalized_at <= datetime.combine(end_date, datetime.max.time()),
                FinalizedDecision.planned_purchase_date <= end_date
            )
        )
    if project_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.project_id.in_(project_ids))
    if supplier_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.supplier_id.in_(supplier_ids))
    
    decisions = decisions_query.all()
    
    # Cash Flow Analysis
    cash_flow_data = {}
    
    # Collect all dates and aggregate cash flows
    for decision in decisions:
        # Outflow (actual payments)
        if decision.actual_payment_confirmed_at:
            payment_date = decision.actual_payment_confirmed_at.date()
            if payment_date not in cash_flow_data:
                cash_flow_data[payment_date] = {'inflow': 0, 'outflow': 0}
            cash_flow_data[payment_date]['outflow'] += float(decision.actual_payment_amount or 0)
    
    # Get cashflow events for inflows
    cashflow_query = db.query(CashflowEvent).filter(
        CashflowEvent.event_type == 'ACTUAL_INFLOW'
    )
    
    if start_date:
        cashflow_query = cashflow_query.filter(CashflowEvent.event_date >= start_date)
    if end_date:
        cashflow_query = cashflow_query.filter(CashflowEvent.event_date <= end_date)
    if project_ids:
        cashflow_query = cashflow_query.filter(CashflowEvent.project_id.in_(project_ids))
    
    cashflow_events = cashflow_query.all()
    
    for event in cashflow_events:
        event_date = event.event_date
        if event_date not in cash_flow_data:
            cash_flow_data[event_date] = {'inflow': 0, 'outflow': 0}
        cash_flow_data[event_date]['inflow'] += float(event.amount or 0)
    
    # Sort by date and calculate cumulative
    sorted_dates = sorted(cash_flow_data.keys())
    dates_str = [d.isoformat() for d in sorted_dates]
    inflows = [cash_flow_data[d]['inflow'] for d in sorted_dates]
    outflows = [cash_flow_data[d]['outflow'] for d in sorted_dates]
    net_flows = [inflows[i] - outflows[i] for i in range(len(inflows))]
    
    cumulative_balance = []
    balance = 0
    for net in net_flows:
        balance += net
        cumulative_balance.append(balance)
    
    # Budget vs Actuals by Project
    budget_vs_actual = []
    
    # Group by project
    project_data = {}
    for decision in decisions:
        project_id = decision.project_id
        if project_id not in project_data:
            project_data[project_id] = {
                'project_name': decision.project.name if decision.project else 'Unknown',
                'planned_cost': 0,
                'actual_cost': 0
            }
        project_data[project_id]['planned_cost'] += float(decision.planned_cost or 0)
        project_data[project_id]['actual_cost'] += float(decision.actual_payment_amount or 0)
    
    total_planned = 0
    total_actual = 0
    
    for project_id, data in project_data.items():
        planned = data['planned_cost']
        actual = data['actual_cost']
        variance_amount = actual - planned
        variance_percent = (variance_amount / planned * 100) if planned > 0 else 0
        
        budget_vs_actual.append({
            'project_name': data['project_name'],
            'planned_cost': round(planned, 2),
            'actual_cost': round(actual, 2),
            'variance_amount': round(variance_amount, 2),
            'variance_percent': round(variance_percent, 2)
        })
        
        total_planned += planned
        total_actual += actual
    
    # Add grand total
    total_variance = total_actual - total_planned
    total_variance_percent = (total_variance / total_planned * 100) if total_planned > 0 else 0
    
    budget_vs_actual.append({
        'project_name': 'GRAND TOTAL',
        'planned_cost': round(total_planned, 2),
        'actual_cost': round(total_actual, 2),
        'variance_amount': round(total_variance, 2),
        'variance_percent': round(total_variance_percent, 2)
    })
    
    return {
        'cash_flow': {
            'dates': dates_str,
            'inflow': [round(x, 2) for x in inflows],
            'outflow': [round(x, 2) for x in outflows],
            'net_flow': [round(x, 2) for x in net_flows],
            'cumulative_balance': [round(x, 2) for x in cumulative_balance]
        },
        'budget_vs_actual': budget_vs_actual
    }


def aggregate_evm_analytics(
    db: Session,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_ids: Optional[List[int]]
) -> Dict[str, Any]:
    """Aggregate EVM (Earned Value Management) analytics data."""
    
    # Base query
    decisions_query = db.query(FinalizedDecision).filter(
        FinalizedDecision.status == 'FINALIZED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.filter(
            or_(
                FinalizedDecision.planned_purchase_date >= start_date,
                FinalizedDecision.actual_payment_confirmed_at >= datetime.combine(start_date, datetime.min.time())
            )
        )
    if end_date:
        decisions_query = decisions_query.filter(
            or_(
                FinalizedDecision.planned_purchase_date <= end_date,
                FinalizedDecision.actual_payment_confirmed_at <= datetime.combine(end_date, datetime.max.time())
            )
        )
    if project_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.project_id.in_(project_ids))
    if supplier_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.supplier_id.in_(supplier_ids))
    
    decisions = decisions_query.all()
    
    # EVM Performance over time
    evm_data = {}
    
    for decision in decisions:
        # Planned Value (PV) - based on planned purchase date
        if decision.planned_purchase_date:
            pv_date = decision.planned_purchase_date
            if pv_date not in evm_data:
                evm_data[pv_date] = {'pv': 0, 'ev': 0, 'ac': 0}
            evm_data[pv_date]['pv'] += float(decision.planned_cost or 0)
        
        # Earned Value (EV) - based on PM acceptance date
        if decision.pm_accepted_at:
            ev_date = decision.pm_accepted_at.date()
            if ev_date not in evm_data:
                evm_data[ev_date] = {'pv': 0, 'ev': 0, 'ac': 0}
            evm_data[ev_date]['ev'] += float(decision.planned_cost or 0)
        
        # Actual Cost (AC) - based on actual payment date
        if decision.actual_payment_confirmed_at:
            ac_date = decision.actual_payment_confirmed_at.date()
            if ac_date not in evm_data:
                evm_data[ac_date] = {'pv': 0, 'ev': 0, 'ac': 0}
            evm_data[ac_date]['ac'] += float(decision.actual_payment_amount or 0)
    
    # Sort and create cumulative values
    sorted_dates = sorted(evm_data.keys())
    dates_str = [d.isoformat() for d in sorted_dates]
    
    cumulative_pv = []
    cumulative_ev = []
    cumulative_ac = []
    pv_sum = ev_sum = ac_sum = 0
    
    for d in sorted_dates:
        pv_sum += evm_data[d]['pv']
        ev_sum += evm_data[d]['ev']
        ac_sum += evm_data[d]['ac']
        cumulative_pv.append(round(pv_sum, 2))
        cumulative_ev.append(round(ev_sum, 2))
        cumulative_ac.append(round(ac_sum, 2))
    
    # KPI Trends (CPI and SPI over time)
    cpi_values = []
    spi_values = []
    
    for i in range(len(cumulative_ev)):
        ev = cumulative_ev[i]
        ac = cumulative_ac[i]
        pv = cumulative_pv[i]
        
        cpi = (ev / ac) if ac > 0 else 1.0
        spi = (ev / pv) if pv > 0 else 1.0
        
        cpi_values.append(round(cpi, 3))
        spi_values.append(round(spi, 3))
    
    # Project KPI Breakdown
    project_kpis = []
    project_data = {}
    
    for decision in decisions:
        project_id = decision.project_id
        if project_id not in project_data:
            project_data[project_id] = {
                'project_name': decision.project.name if decision.project else 'Unknown',
                'pv': 0,
                'ev': 0,
                'ac': 0
            }
        
        project_data[project_id]['pv'] += float(decision.planned_cost or 0)
        
        if decision.pm_accepted_at:
            project_data[project_id]['ev'] += float(decision.planned_cost or 0)
        
        if decision.actual_payment_confirmed_at:
            project_data[project_id]['ac'] += float(decision.actual_payment_amount or 0)
    
    for project_id, data in project_data.items():
        pv = data['pv']
        ev = data['ev']
        ac = data['ac']
        
        sv = ev - pv  # Schedule Variance
        cv = ev - ac  # Cost Variance
        spi = (ev / pv) if pv > 0 else 1.0
        cpi = (ev / ac) if ac > 0 else 1.0
        
        # Estimate at Completion (EAC) = AC + (BAC - EV) / CPI
        bac = pv  # Budget at Completion
        eac = ac + ((bac - ev) / cpi) if cpi > 0 else bac
        
        # Estimate to Complete (ETC) = EAC - AC
        etc = eac - ac
        
        project_kpis.append({
            'project_name': data['project_name'],
            'pv': round(pv, 2),
            'ev': round(ev, 2),
            'ac': round(ac, 2),
            'sv': round(sv, 2),
            'cv': round(cv, 2),
            'spi': round(spi, 3),
            'cpi': round(cpi, 3),
            'eac': round(eac, 2),
            'etc': round(etc, 2)
        })
    
    return {
        'evm_performance': {
            'dates': dates_str,
            'pv': cumulative_pv,
            'ev': cumulative_ev,
            'ac': cumulative_ac
        },
        'kpi_trends': {
            'dates': dates_str,
            'cpi': cpi_values,
            'spi': spi_values
        },
        'project_kpis': project_kpis
    }


def aggregate_risk_forecasts(
    db: Session,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_ids: Optional[List[int]]
) -> Dict[str, Any]:
    """Aggregate risk and forecast data."""
    
    # Base query
    decisions_query = db.query(FinalizedDecision).filter(
        FinalizedDecision.status == 'FINALIZED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.filter(
            FinalizedDecision.planned_purchase_date >= start_date
        )
    if end_date:
        decisions_query = decisions_query.filter(
            FinalizedDecision.planned_purchase_date <= end_date
        )
    if project_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.project_id.in_(project_ids))
    if supplier_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.supplier_id.in_(supplier_ids))
    
    decisions = decisions_query.all()
    
    # Calculate payment delays
    payment_delays = []
    for decision in decisions:
        if decision.actual_payment_confirmed_at and decision.planned_purchase_date:
            actual_date = decision.actual_payment_confirmed_at.date()
            planned_date = decision.planned_purchase_date
            delay_days = (actual_date - planned_date).days
            payment_delays.append(delay_days)
    
    # Calculate P50 and P90
    p50_delay = calculate_percentile(payment_delays, 50) if payment_delays else 0
    p90_delay = calculate_percentile(payment_delays, 90) if payment_delays else 0
    
    # Payment delay distribution (histogram data)
    delay_histogram = {}
    for delay in payment_delays:
        # Bucket delays into 10-day intervals
        bucket = (delay // 10) * 10
        delay_histogram[bucket] = delay_histogram.get(bucket, 0) + 1
    
    histogram_data = [
        {'delay_bucket': k, 'count': v}
        for k, v in sorted(delay_histogram.items())
    ]
    
    # Top 5 Highest Risk Items
    risk_items = []
    for decision in decisions:
        planned_cost = float(decision.planned_cost or 0)
        actual_cost = float(decision.actual_payment_amount or 0)
        cost_variance = actual_cost - planned_cost
        
        schedule_delay = 0
        if decision.actual_payment_confirmed_at and decision.planned_purchase_date:
            schedule_delay = (decision.actual_payment_confirmed_at.date() - decision.planned_purchase_date).days
        
        # Calculate risk score (combination of cost overrun and schedule delay)
        risk_score = abs(cost_variance) + (abs(schedule_delay) * 10)  # Weight schedule delay more
        
        risk_items.append({
            'item_name': decision.item.name if decision.item else decision.item_code,
            'project_name': decision.project.name if decision.project else 'Unknown',
            'cost_variance': round(cost_variance, 2),
            'schedule_delay': schedule_delay,
            'risk_score': risk_score
        })
    
    # Sort by risk score and take top 5
    risk_items.sort(key=lambda x: x['risk_score'], reverse=True)
    top_risk_items = risk_items[:5]
    
    return {
        'delay_forecast': {
            'p50': round(p50_delay, 1),
            'p90': round(p90_delay, 1)
        },
        'payment_delay_histogram': histogram_data,
        'top_risk_items': top_risk_items
    }


def aggregate_operational_performance(
    db: Session,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_ids: Optional[List[int]]
) -> Dict[str, Any]:
    """Aggregate operational performance data."""
    
    # Base query
    decisions_query = db.query(FinalizedDecision).filter(
        FinalizedDecision.status == 'FINALIZED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.filter(
            FinalizedDecision.decision_finalized_at >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        decisions_query = decisions_query.filter(
            FinalizedDecision.decision_finalized_at <= datetime.combine(end_date, datetime.max.time())
        )
    if project_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.project_id.in_(project_ids))
    if supplier_ids:
        decisions_query = decisions_query.filter(FinalizedDecision.supplier_id.in_(supplier_ids))
    
    decisions = decisions_query.all()
    
    # Supplier Scorecard
    supplier_data = {}
    
    for decision in decisions:
        supplier_id = decision.supplier_id
        if not supplier_id:
            continue
        
        if supplier_id not in supplier_data:
            supplier_data[supplier_id] = {
                'supplier_name': decision.supplier.name if decision.supplier else 'Unknown',
                'total_orders': 0,
                'on_time_deliveries': 0,
                'total_planned_cost': 0,
                'total_actual_cost': 0
            }
        
        supplier_data[supplier_id]['total_orders'] += 1
        supplier_data[supplier_id]['total_planned_cost'] += float(decision.planned_cost or 0)
        supplier_data[supplier_id]['total_actual_cost'] += float(decision.actual_payment_amount or 0)
        
        # Check on-time delivery
        if decision.actual_delivery_date and decision.planned_delivery_date:
            if decision.actual_delivery_date <= decision.planned_delivery_date:
                supplier_data[supplier_id]['on_time_deliveries'] += 1
    
    supplier_scorecard = []
    for supplier_id, data in supplier_data.items():
        total_orders = data['total_orders']
        on_time_rate = (data['on_time_deliveries'] / total_orders * 100) if total_orders > 0 else 0
        
        planned = data['total_planned_cost']
        actual = data['total_actual_cost']
        avg_cost_variance = ((actual - planned) / planned * 100) if planned > 0 else 0
        
        supplier_scorecard.append({
            'supplier_name': data['supplier_name'],
            'total_orders': total_orders,
            'on_time_delivery_rate': round(on_time_rate, 2),
            'avg_cost_variance_percent': round(avg_cost_variance, 2)
        })
    
    # Procurement Cycle Time
    cycle_times = []
    for decision in decisions:
        if decision.decision_finalized_at and decision.pm_accepted_at:
            cycle_time = (decision.pm_accepted_at - decision.decision_finalized_at).days
            cycle_times.append(cycle_time)
    
    # Create histogram for cycle times
    cycle_time_histogram = {}
    for cycle_time in cycle_times:
        # Bucket into 5-day intervals
        bucket = (cycle_time // 5) * 5
        cycle_time_histogram[bucket] = cycle_time_histogram.get(bucket, 0) + 1
    
    cycle_time_data = [
        {'cycle_time_bucket': k, 'count': v}
        for k, v in sorted(cycle_time_histogram.items())
    ]
    
    return {
        'supplier_scorecard': supplier_scorecard,
        'procurement_cycle_time': cycle_time_data
    }


@router.get("/")
async def get_reports_data(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    project_ids: Optional[str] = Query(None, description="Comma-separated project IDs"),
    supplier_ids: Optional[str] = Query(None, description="Comma-separated supplier IDs"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get comprehensive reports data for all tabs.
    This is the main data aggregation endpoint.
    """
    
    # Parse comma-separated IDs
    project_id_list = [int(x) for x in project_ids.split(',')] if project_ids else None
    supplier_id_list = [int(x) for x in supplier_ids.split(',')] if supplier_ids else None
    
    # For PMs, filter to only their assigned projects
    if current_user.role == 'pm':
        user_project_ids = [p.id for p in current_user.assigned_projects]
        if project_id_list:
            # Intersect with user's projects
            project_id_list = [pid for pid in project_id_list if pid in user_project_ids]
        else:
            project_id_list = user_project_ids
    
    # Aggregate all data
    financial_summary = aggregate_financial_summary(db, start_date, end_date, project_id_list, supplier_id_list)
    evm_analytics = aggregate_evm_analytics(db, start_date, end_date, project_id_list, supplier_id_list)
    risk_forecasts = aggregate_risk_forecasts(db, start_date, end_date, project_id_list, supplier_id_list)
    operational_performance = aggregate_operational_performance(db, start_date, end_date, project_id_list, supplier_id_list)
    
    return {
        'financial_summary': financial_summary,
        'evm_analytics': evm_analytics,
        'risk_forecasts': risk_forecasts,
        'operational_performance': operational_performance
    }


@router.get("/export/excel")
async def export_reports_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    project_ids: Optional[str] = Query(None),
    supplier_ids: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export reports data to Excel file."""
    
    # Get the same data as the main endpoint
    project_id_list = [int(x) for x in project_ids.split(',')] if project_ids else None
    supplier_id_list = [int(x) for x in supplier_ids.split(',')] if supplier_ids else None
    
    # For PMs, filter to only their assigned projects
    if current_user.role == 'pm':
        user_project_ids = [p.id for p in current_user.assigned_projects]
        if project_id_list:
            project_id_list = [pid for pid in project_id_list if pid in user_project_ids]
        else:
            project_id_list = user_project_ids
    
    financial_summary = aggregate_financial_summary(db, start_date, end_date, project_id_list, supplier_id_list)
    evm_analytics = aggregate_evm_analytics(db, start_date, end_date, project_id_list, supplier_id_list)
    risk_forecasts = aggregate_risk_forecasts(db, start_date, end_date, project_id_list, supplier_id_list)
    operational_performance = aggregate_operational_performance(db, start_date, end_date, project_id_list, supplier_id_list)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Financial Summary
    ws1 = wb.active
    ws1.title = "Financial Summary"
    
    # Budget vs Actuals Table
    ws1.append(["Budget vs Actuals Summary"])
    ws1.append(["Project Name", "Planned Cost", "Actual Cost", "Variance ($)", "Variance (%)"])
    
    for cell in ws1[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in financial_summary['budget_vs_actual']:
        ws1.append([
            row['project_name'],
            row['planned_cost'],
            row['actual_cost'],
            row['variance_amount'],
            row['variance_percent']
        ])
    
    # Sheet 2: EVM Analytics
    ws2 = wb.create_sheet("EVM Analytics")
    ws2.append(["Project KPI Breakdown"])
    ws2.append(["Project", "PV", "EV", "AC", "SV ($)", "CV ($)", "SPI", "CPI", "EAC", "ETC"])
    
    for cell in ws2[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in evm_analytics['project_kpis']:
        ws2.append([
            row['project_name'],
            row['pv'],
            row['ev'],
            row['ac'],
            row['sv'],
            row['cv'],
            row['spi'],
            row['cpi'],
            row['eac'],
            row['etc']
        ])
    
    # Sheet 3: Risk & Forecasts
    ws3 = wb.create_sheet("Risk & Forecasts")
    ws3.append(["Delay Forecast"])
    ws3.append(["Metric", "Value (Days)"])
    ws3.append(["P50 (Median)", risk_forecasts['delay_forecast']['p50']])
    ws3.append(["P90 (90th Percentile)", risk_forecasts['delay_forecast']['p90']])
    ws3.append([])
    
    ws3.append(["Top Risk Items"])
    ws3.append(["Item Name", "Project", "Cost Variance ($)", "Schedule Delay (Days)"])
    
    for cell in ws3[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in risk_forecasts['top_risk_items']:
        ws3.append([
            row['item_name'],
            row['project_name'],
            row['cost_variance'],
            row['schedule_delay']
        ])
    
    # Sheet 4: Operational Performance
    ws4 = wb.create_sheet("Operational Performance")
    ws4.append(["Supplier Scorecard"])
    ws4.append(["Supplier Name", "Total Orders", "On-Time Delivery Rate (%)", "Avg Cost Variance (%)"])
    
    for cell in ws4[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in operational_performance['supplier_scorecard']:
        ws4.append([
            row['supplier_name'],
            row['total_orders'],
            row['on_time_delivery_rate'],
            row['avg_cost_variance_percent']
        ])
    
    # Auto-size columns
    for ws in [ws1, ws2, ws3, ws4]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/filters/projects")
async def get_projects_for_filter(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get list of projects for filter dropdown."""
    if current_user.role == 'pm':
        projects = current_user.assigned_projects
    else:
        projects = db.query(Project).filter(Project.is_active == True).all()
    
    return [{'id': p.id, 'name': p.name, 'code': p.code} for p in projects]


@router.get("/filters/suppliers")
async def get_suppliers_for_filter(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get list of suppliers for filter dropdown."""
    suppliers = db.query(Supplier).filter(Supplier.is_active == True).all()
    return [{'id': s.id, 'name': s.name} for s in suppliers]

