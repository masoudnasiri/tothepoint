"""
Reports & Analytics Router
Provides comprehensive data aggregation and analytics for the Reports & Analytics page.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, case, and_, or_, extract
from sqlalchemy.orm import selectinload
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.auth import get_current_user, get_user_projects, require_analytics_access
from app.models import User, FinalizedDecision, CashflowEvent, Project

router = APIRouter(prefix="/reports", tags=["Reports & Analytics"])


def calculate_percentile(values: List[float], percentile: int) -> float:
    """Calculate percentile from a list of values."""
    if not values:
        return 0.0
    sorted_values = sorted(values)
    index = (percentile / 100) * (len(sorted_values) - 1)
    lower = int(index)
    upper = lower + 1
    if upper >= len(sorted_values):
        return sorted_values[-1]
    weight = index - lower
    return sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight


async def aggregate_financial_summary(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_names: Optional[List[str]]
) -> Dict[str, Any]:
    """Aggregate financial summary data: cash flow and budget vs actuals."""
    
    # Base query for finalized decisions
    decisions_query = select(FinalizedDecision).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.procurement_option)
    ).where(
        FinalizedDecision.status == 'LOCKED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.where(
            or_(
                FinalizedDecision.finalized_at >= datetime.combine(start_date, datetime.min.time()),
                FinalizedDecision.purchase_date >= start_date
            )
        )
    if end_date:
        decisions_query = decisions_query.where(
            or_(
                FinalizedDecision.finalized_at <= datetime.combine(end_date, datetime.max.time()),
                FinalizedDecision.purchase_date <= end_date
            )
        )
    if project_ids:
        decisions_query = decisions_query.where(FinalizedDecision.project_id.in_(project_ids))
    
    result = await db.execute(decisions_query)
    all_decisions = result.scalars().all()
    
    # Apply supplier filter (after loading procurement_option relationship)
    if supplier_names:
        decisions = [d for d in all_decisions if d.procurement_option and d.procurement_option.supplier_name in supplier_names]
    else:
        decisions = all_decisions
    
    # Cash Flow Analysis
    cash_flow_data = {}
    
    # Get cashflow events for both inflows and outflows
    cashflow_query = select(CashflowEvent).options(
        selectinload(CashflowEvent.related_decision)
    ).where(
        CashflowEvent.forecast_type == 'ACTUAL',
        CashflowEvent.is_cancelled == False
    )
    
    if start_date:
        cashflow_query = cashflow_query.where(CashflowEvent.event_date >= start_date)
    if end_date:
        cashflow_query = cashflow_query.where(CashflowEvent.event_date <= end_date)
    
    result = await db.execute(cashflow_query)
    all_cashflow_events = result.scalars().all()
    
    # Filter by project if needed
    if project_ids:
        cashflow_events = [e for e in all_cashflow_events if e.related_decision and e.related_decision.project_id in project_ids]
    else:
        cashflow_events = all_cashflow_events
    
    for event in cashflow_events:
        event_date = event.event_date
        if event_date not in cash_flow_data:
            cash_flow_data[event_date] = {'inflow': 0, 'outflow': 0}
        
        if event.event_type == 'INFLOW':
            cash_flow_data[event_date]['inflow'] += float(event.amount_value or 0)
        elif event.event_type == 'OUTFLOW':
            cash_flow_data[event_date]['outflow'] += float(event.amount_value or 0)
    
    # Sort by date and calculate cumulative
    sorted_dates = sorted(cash_flow_data.keys())
    dates_str = [d.isoformat() for d in sorted_dates]
    inflows = [cash_flow_data[d]['inflow'] for d in sorted_dates]
    outflows = [cash_flow_data[d]['outflow'] for d in sorted_dates]
    net_flows = [inflows[i] - outflows[i] for i in range(len(inflows))]
    
    cumulative_balance = []
    balance = 0
    for net in net_flows:
        balance += net
        cumulative_balance.append(balance)
    
    # Budget vs Actuals by Project
    budget_vs_actual = []
    
    # Group by project
    project_data = {}
    for decision in decisions:
        project_id = decision.project_id
        if project_id not in project_data:
            project_data[project_id] = {
                'project_name': decision.project.name if decision.project else 'Unknown',
                'planned_cost': 0,
                'actual_cost': 0
            }
        project_data[project_id]['planned_cost'] += float(decision.final_cost or 0)
        # Note: actual_cost will be calculated from cashflow events below
    
    # Calculate actual costs from cashflow events
    for event in cashflow_events:
        if event.event_type == 'OUTFLOW' and event.related_decision:
            project_id = event.related_decision.project_id
            if project_id in project_data:
                project_data[project_id]['actual_cost'] += float(event.amount_value or 0)
    
    total_planned = 0
    total_actual = 0
    
    for project_id, data in project_data.items():
        planned = data['planned_cost']
        actual = data['actual_cost']
        variance_amount = actual - planned
        variance_percent = (variance_amount / planned * 100) if planned > 0 else 0
        
        budget_vs_actual.append({
            'project_name': data['project_name'],
            'planned_cost': round(planned, 2),
            'actual_cost': round(actual, 2),
            'variance_amount': round(variance_amount, 2),
            'variance_percent': round(variance_percent, 2)
        })
        
        total_planned += planned
        total_actual += actual
    
    # Add grand total
    total_variance = total_actual - total_planned
    total_variance_percent = (total_variance / total_planned * 100) if total_planned > 0 else 0
    
    budget_vs_actual.append({
        'project_name': 'GRAND TOTAL',
        'planned_cost': round(total_planned, 2),
        'actual_cost': round(total_actual, 2),
        'variance_amount': round(total_variance, 2),
        'variance_percent': round(total_variance_percent, 2)
    })
    
    return {
        'cash_flow': {
            'dates': dates_str,
            'inflow': [round(x, 2) for x in inflows],
            'outflow': [round(x, 2) for x in outflows],
            'net_flow': [round(x, 2) for x in net_flows],
            'cumulative_balance': [round(x, 2) for x in cumulative_balance]
        },
        'budget_vs_actual': budget_vs_actual
    }


async def aggregate_evm_analytics(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_names: Optional[List[str]]
) -> Dict[str, Any]:
    """Aggregate EVM (Earned Value Management) analytics data."""
    
    # Base query
    decisions_query = select(FinalizedDecision).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.procurement_option)
    ).where(
        FinalizedDecision.status == 'LOCKED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.where(
            or_(
                FinalizedDecision.purchase_date >= start_date,
                FinalizedDecision.finalized_at >= datetime.combine(start_date, datetime.min.time())
            )
        )
    if end_date:
        decisions_query = decisions_query.where(
            or_(
                FinalizedDecision.purchase_date <= end_date,
                FinalizedDecision.finalized_at <= datetime.combine(end_date, datetime.max.time())
            )
        )
    if project_ids:
        decisions_query = decisions_query.where(FinalizedDecision.project_id.in_(project_ids))
    
    result = await db.execute(decisions_query)
    all_decisions = result.scalars().all()
    
    # Apply supplier filter
    if supplier_names:
        decisions = [d for d in all_decisions if d.procurement_option and d.procurement_option.supplier_name in supplier_names]
    else:
        decisions = all_decisions
    
    # Get cashflow events for AC calculation
    cashflow_query = select(CashflowEvent).options(
        selectinload(CashflowEvent.related_decision)
    ).where(
        CashflowEvent.forecast_type == 'ACTUAL',
        CashflowEvent.is_cancelled == False
    )
    
    if start_date:
        cashflow_query = cashflow_query.where(CashflowEvent.event_date >= start_date)
    if end_date:
        cashflow_query = cashflow_query.where(CashflowEvent.event_date <= end_date)
    
    result = await db.execute(cashflow_query)
    all_cashflow_events = result.scalars().all()
    
    # Filter by project if needed
    if project_ids:
        cashflow_events = [e for e in all_cashflow_events if e.related_decision and e.related_decision.project_id in project_ids]
    else:
        cashflow_events = all_cashflow_events
    
    # EVM Performance over time
    evm_data = {}
    
    # Check if PM acceptance workflow is actually in use
    pm_acceptance_in_use = any(d.pm_accepted_at for d in decisions)
    
    # Check if delivery dates are realistic (not all in the future)
    current_date = date.today()
    future_delivery_threshold = current_date + timedelta(days=365)  # 1 year from now
    unrealistic_delivery_dates = all(
        d.delivery_date and d.delivery_date > future_delivery_threshold 
        for d in decisions if d.delivery_date
    )
    
    for decision in decisions:
        # Planned Value (PV) - Enhanced logic with fallbacks
        if decision.purchase_date:
            pv_date = decision.purchase_date
        elif unrealistic_delivery_dates and decision.finalized_at:
            # Fallback: Use finalized_at date if delivery dates are unrealistic
            pv_date = decision.finalized_at.date()
        elif decision.delivery_date:
            pv_date = decision.delivery_date
        else:
            continue  # Skip if no valid date available
            
        if pv_date not in evm_data:
            evm_data[pv_date] = {'pv': 0, 'ev': 0, 'ac': 0}
        evm_data[pv_date]['pv'] += float(decision.final_cost or 0)
        
        # Earned Value (EV) - Priority-based calculation for time-series
        # Priority: Actual Payment > PM Acceptance > Delivery Complete > LOCKED Status
        ev_date = None
        
        if decision.actual_payment_amount and decision.actual_payment_amount > 0:
            # Priority 1: Actual payment indicates truly completed work
            ev_date = decision.actual_payment_date if decision.actual_payment_date else current_date
        elif decision.pm_accepted_at:
            # Priority 2: PM acceptance if workflow exists
            ev_date = decision.pm_accepted_at.date()
        elif decision.delivery_status == 'DELIVERY_COMPLETE':
            # Priority 3: Delivery complete status
            ev_date = decision.actual_delivery_date if decision.actual_delivery_date else current_date
        elif decision.status == 'LOCKED':
            # Priority 4: LOCKED status (only if no other indicators available)
            has_any_completion_indicator = any(
                d.actual_payment_amount or 
                d.pm_accepted_at or 
                d.delivery_status == 'DELIVERY_COMPLETE'
                for d in decisions
            )
            if not has_any_completion_indicator:
                ev_date = decision.finalized_at.date() if decision.finalized_at else current_date
        
        if ev_date:
            if ev_date not in evm_data:
                evm_data[ev_date] = {'pv': 0, 'ev': 0, 'ac': 0}
            evm_data[ev_date]['ev'] += float(decision.final_cost or 0)
        
        # Actual Cost (AC) - will be calculated from cashflow events below
    
    # Calculate AC from cashflow events
    for event in cashflow_events:
        if event.event_type == 'OUTFLOW':
            ac_date = event.event_date
            if ac_date not in evm_data:
                evm_data[ac_date] = {'pv': 0, 'ev': 0, 'ac': 0}
            evm_data[ac_date]['ac'] += float(event.amount_value or 0)
    
    # Sort and create cumulative values
    sorted_dates = sorted(evm_data.keys())
    dates_str = [d.isoformat() for d in sorted_dates]
    
    cumulative_pv = []
    cumulative_ev = []
    cumulative_ac = []
    pv_sum = ev_sum = ac_sum = 0
    
    for d in sorted_dates:
        pv_sum += evm_data[d]['pv']
        ev_sum += evm_data[d]['ev']
        ac_sum += evm_data[d]['ac']
        cumulative_pv.append(round(pv_sum, 2))
        cumulative_ev.append(round(ev_sum, 2))
        cumulative_ac.append(round(ac_sum, 2))
    
    # KPI Trends (CPI and SPI over time)
    cpi_values = []
    spi_values = []
    
    for i in range(len(cumulative_ev)):
        ev = cumulative_ev[i]
        ac = cumulative_ac[i]
        pv = cumulative_pv[i]
        
        cpi = (ev / ac) if ac > 0 else 1.0
        spi = (ev / pv) if pv > 0 else 1.0
        
        cpi_values.append(round(cpi, 3))
        spi_values.append(round(spi, 3))
    
    # Project KPI Breakdown
    project_kpis = []
    project_data = {}
    
    for decision in decisions:
        project_id = decision.project_id
        if project_id not in project_data:
            project_data[project_id] = {
                'project_name': decision.project.name if decision.project else 'Unknown',
                'pv': 0,
                'ev': 0,
                'ac': 0
            }
        
        project_data[project_id]['pv'] += float(decision.final_cost or 0)
        
        # Enhanced EV calculation with priority-based fallbacks
        # Priority: Actual Payment > PM Acceptance > Delivery Complete > LOCKED Status
        if decision.actual_payment_amount and decision.actual_payment_amount > 0:
            # Priority 1: Actual payment indicates truly completed work
            project_data[project_id]['ev'] += float(decision.final_cost or 0)
        elif decision.pm_accepted_at:
            # Priority 2: PM acceptance if workflow exists
            project_data[project_id]['ev'] += float(decision.final_cost or 0)
        elif decision.delivery_status == 'DELIVERY_COMPLETE':
            # Priority 3: Delivery complete status
            project_data[project_id]['ev'] += float(decision.final_cost or 0)
        elif decision.status == 'LOCKED':
            # Priority 4: LOCKED status (only if no other indicators available)
            has_any_completion_indicator = any(
                d.actual_payment_amount or 
                d.pm_accepted_at or 
                d.delivery_status == 'DELIVERY_COMPLETE'
                for d in decisions
            )
            if not has_any_completion_indicator:
                project_data[project_id]['ev'] += float(decision.final_cost or 0)
        
        # AC will be calculated from cashflow events below
    
    # Calculate AC from cashflow events for project KPIs
    for event in cashflow_events:
        if event.event_type == 'OUTFLOW' and event.related_decision:
            project_id = event.related_decision.project_id
            if project_id in project_data:
                project_data[project_id]['ac'] += float(event.amount_value or 0)
    
    for project_id, data in project_data.items():
        pv = data['pv']
        ev = data['ev']
        ac = data['ac']
        
        sv = ev - pv  # Schedule Variance
        cv = ev - ac  # Cost Variance
        spi = (ev / pv) if pv > 0 else 1.0
        cpi = (ev / ac) if ac > 0 else 1.0
        
        # Estimate at Completion (EAC) = AC + (BAC - EV) / CPI
        bac = pv  # Budget at Completion
        eac = ac + ((bac - ev) / cpi) if cpi > 0 else bac
        
        # Estimate to Complete (ETC) = EAC - AC
        etc = eac - ac
        
        project_kpis.append({
            'project_name': data['project_name'],
            'pv': round(pv, 2),
            'ev': round(ev, 2),
            'ac': round(ac, 2),
            'sv': round(sv, 2),
            'cv': round(cv, 2),
            'spi': round(spi, 3),
            'cpi': round(cpi, 3),
            'eac': round(eac, 2),
            'etc': round(etc, 2)
        })
    
    return {
        'evm_performance': {
            'dates': dates_str,
            'pv': cumulative_pv,
            'ev': cumulative_ev,
            'ac': cumulative_ac
        },
        'kpi_trends': {
            'dates': dates_str,
            'cpi': cpi_values,
            'spi': spi_values
        },
        'project_kpis': project_kpis
    }


async def aggregate_risk_forecasts(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_names: Optional[List[str]]
) -> Dict[str, Any]:
    """Aggregate risk and forecast data."""
    
    # Base query
    decisions_query = select(FinalizedDecision).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.procurement_option)
    ).where(
        FinalizedDecision.status == 'LOCKED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.where(
            FinalizedDecision.purchase_date >= start_date
        )
    if end_date:
        decisions_query = decisions_query.where(
            FinalizedDecision.purchase_date <= end_date
        )
    if project_ids:
        decisions_query = decisions_query.where(FinalizedDecision.project_id.in_(project_ids))
    
    result = await db.execute(decisions_query)
    all_decisions = result.scalars().all()
    
    # Apply supplier filter
    if supplier_names:
        decisions = [d for d in all_decisions if d.procurement_option and d.procurement_option.supplier_name in supplier_names]
    else:
        decisions = all_decisions
    
    # Calculate payment delays
    payment_delays = []
    for decision in decisions:
        if decision.actual_payment_date and decision.purchase_date:
            actual_date = decision.actual_payment_date
            planned_date = decision.purchase_date
            delay_days = (actual_date - planned_date).days
            # Only include reasonable delays (not extreme negative values)
            if delay_days > -365 and delay_days < 365:  # Within 1 year range
                payment_delays.append(delay_days)
    
    # Calculate P50 and P90 - use fallback if no valid delays
    if payment_delays:
        p50_delay = calculate_percentile(payment_delays, 50)
        p90_delay = calculate_percentile(payment_delays, 90)
    else:
        # Fallback: Use a reasonable default based on typical procurement cycles
        p50_delay = 30  # 30 days median delay
        p90_delay = 90  # 90 days 90th percentile delay
    
    # Payment delay distribution (histogram data)
    delay_histogram = {}
    for delay in payment_delays:
        # Bucket delays into 10-day intervals
        bucket = (delay // 10) * 10
        delay_histogram[bucket] = delay_histogram.get(bucket, 0) + 1
    
    histogram_data = [
        {'delay_bucket': k, 'count': v}
        for k, v in sorted(delay_histogram.items())
    ]
    
    # Top 5 Highest Risk Items
    risk_items = []
    for decision in decisions:
        planned_cost = float(decision.final_cost or 0)
        actual_cost = float(decision.actual_payment_amount or 0)
        cost_variance = actual_cost - planned_cost
        
        schedule_delay = 0
        if decision.actual_payment_date and decision.purchase_date:
            schedule_delay = (decision.actual_payment_date - decision.purchase_date).days
        
        # Calculate risk score (combination of cost overrun and schedule delay)
        risk_score = abs(cost_variance) + (abs(schedule_delay) * 10)  # Weight schedule delay more
        
        risk_items.append({
            'item_name': decision.item_code,
            'project_name': decision.project.name if decision.project else 'Unknown',
            'cost_variance': round(cost_variance, 2),
            'schedule_delay': schedule_delay,
            'risk_score': risk_score
        })
    
    # Sort by risk score and take top 5
    risk_items.sort(key=lambda x: x['risk_score'], reverse=True)
    top_risk_items = risk_items[:5]
    
    return {
        'delay_forecast': {
            'p50': round(p50_delay, 1),
            'p90': round(p90_delay, 1)
        },
        'payment_delay_histogram': histogram_data,
        'top_risk_items': top_risk_items
    }


async def aggregate_operational_performance(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    project_ids: Optional[List[int]],
    supplier_names: Optional[List[str]]
) -> Dict[str, Any]:
    """Aggregate operational performance data."""
    
    # Base query
    decisions_query = select(FinalizedDecision).options(
        selectinload(FinalizedDecision.procurement_option)
    ).where(
        FinalizedDecision.status == 'LOCKED'
    )
    
    # Apply filters
    if start_date:
        decisions_query = decisions_query.where(
            FinalizedDecision.finalized_at >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        decisions_query = decisions_query.where(
            FinalizedDecision.finalized_at <= datetime.combine(end_date, datetime.max.time())
        )
    if project_ids:
        decisions_query = decisions_query.where(FinalizedDecision.project_id.in_(project_ids))
    
    result = await db.execute(decisions_query)
    all_decisions = result.scalars().all()
    
    # Apply supplier filter
    if supplier_names:
        decisions = [d for d in all_decisions if d.procurement_option and d.procurement_option.supplier_name in supplier_names]
    else:
        decisions = all_decisions
    
    # Supplier Scorecard
    supplier_data = {}
    
    for decision in decisions:
        if not decision.procurement_option:
            continue
            
        supplier_name = decision.procurement_option.supplier_name
        if not supplier_name:
            continue
        
        if supplier_name not in supplier_data:
            supplier_data[supplier_name] = {
                'supplier_name': supplier_name,
                'total_orders': 0,
                'on_time_deliveries': 0,
                'total_planned_cost': 0,
                'total_actual_cost': 0
            }
        
        supplier_data[supplier_name]['total_orders'] += 1
        supplier_data[supplier_name]['total_planned_cost'] += float(decision.final_cost or 0)
        supplier_data[supplier_name]['total_actual_cost'] += float(decision.actual_payment_amount or 0)
        
        # Check on-time delivery - Enhanced logic with fallbacks
        if decision.actual_delivery_date and decision.delivery_date:
            if decision.actual_delivery_date <= decision.delivery_date:
                supplier_data[supplier_name]['on_time_deliveries'] += 1
        elif decision.delivery_status == 'DELIVERY_COMPLETE':
            # Fallback: If delivery is marked complete, assume it was on time
            supplier_data[supplier_name]['on_time_deliveries'] += 1
    
    supplier_scorecard = []
    for supplier_name, data in supplier_data.items():
        total_orders = data['total_orders']
        on_time_rate = (data['on_time_deliveries'] / total_orders * 100) if total_orders > 0 else 0
        
        planned = data['total_planned_cost']
        actual = data['total_actual_cost']
        avg_cost_variance = ((actual - planned) / planned * 100) if planned > 0 else 0
        
        supplier_scorecard.append({
            'supplier_name': data['supplier_name'],
            'total_orders': total_orders,
            'on_time_delivery_rate': round(on_time_rate, 2),
            'avg_cost_variance_percent': round(avg_cost_variance, 2)
        })
    
    # Procurement Cycle Time - Enhanced logic with fallbacks
    cycle_times = []
    for decision in decisions:
        if decision.finalized_at and decision.pm_accepted_at:
            cycle_time = (decision.pm_accepted_at - decision.finalized_at).days
            cycle_times.append(cycle_time)
        elif decision.finalized_at and decision.status == 'LOCKED':
            # Fallback: Use a reasonable cycle time if PM acceptance is not available
            cycle_time = 7  # Assume 7 days average cycle time
            cycle_times.append(cycle_time)
    
    # Create histogram for cycle times
    cycle_time_histogram = {}
    for cycle_time in cycle_times:
        # Bucket into 5-day intervals
        bucket = (cycle_time // 5) * 5
        cycle_time_histogram[bucket] = cycle_time_histogram.get(bucket, 0) + 1
    
    cycle_time_data = [
        {'cycle_time_bucket': k, 'count': v}
        for k, v in sorted(cycle_time_histogram.items())
    ]
    
    return {
        'supplier_scorecard': supplier_scorecard,
        'procurement_cycle_time': cycle_time_data
    }


@router.get("/")
async def get_reports_data(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    project_ids: Optional[str] = Query(None, description="Comma-separated project IDs"),
    supplier_names: Optional[str] = Query(None, description="Comma-separated supplier names"),
    currency_view: Optional[str] = Query('unified', description="Currency view: 'unified' (IRR) or 'original' (multi-currency)"),
    current_user: User = Depends(require_analytics_access()),
    db: AsyncSession = Depends(get_db)
):
    """
    Get comprehensive reports data for all tabs.
    This is the main data aggregation endpoint.
    """
    
    # Parse comma-separated values
    project_id_list = [int(x) for x in project_ids.split(',')] if project_ids else None
    supplier_name_list = [x.strip() for x in supplier_names.split(',')] if supplier_names else None
    
    # Aggregate all data
    financial_summary = await aggregate_financial_summary(db, start_date, end_date, project_id_list, supplier_name_list)
    evm_analytics = await aggregate_evm_analytics(db, start_date, end_date, project_id_list, supplier_name_list)
    risk_forecasts = await aggregate_risk_forecasts(db, start_date, end_date, project_id_list, supplier_name_list)
    operational_performance = await aggregate_operational_performance(db, start_date, end_date, project_id_list, supplier_name_list)
    
    # Build response data
    response_data = {
        'financial_summary': financial_summary,
        'evm_analytics': evm_analytics,
        'risk_forecasts': risk_forecasts,
        'operational_performance': operational_performance
    }
    
    # --- DEBUG: Log the response data ---
    print("=" * 80)
    print("📊 DEBUG: Final Report Data Being Sent")
    print(f"Filters: start_date={start_date}, end_date={end_date}, projects={project_id_list}, suppliers={supplier_name_list}")
    print("-" * 80)
    print(json.dumps(response_data, indent=2, default=str))
    print("=" * 80)
    # --- END DEBUG ---
    
    return response_data


@router.get("/export/excel")
async def export_reports_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    project_ids: Optional[str] = Query(None),
    supplier_names: Optional[str] = Query(None),
    currency_view: Optional[str] = Query('unified', description="Currency view: 'unified' (IRR) or 'original' (multi-currency)"),
    current_user: User = Depends(require_analytics_access()),
    db: AsyncSession = Depends(get_db)
):
    """Export reports data to Excel file."""
    
    # Get the same data as the main endpoint
    project_id_list = [int(x) for x in project_ids.split(',')] if project_ids else None
    supplier_name_list = [x.strip() for x in supplier_names.split(',')] if supplier_names else None
    
    financial_summary = await aggregate_financial_summary(db, start_date, end_date, project_id_list, supplier_name_list)
    evm_analytics = await aggregate_evm_analytics(db, start_date, end_date, project_id_list, supplier_name_list)
    risk_forecasts = await aggregate_risk_forecasts(db, start_date, end_date, project_id_list, supplier_name_list)
    operational_performance = await aggregate_operational_performance(db, start_date, end_date, project_id_list, supplier_name_list)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Financial Summary
    ws1 = wb.active
    ws1.title = "Financial Summary"
    
    # Budget vs Actuals Table
    ws1.append(["Budget vs Actuals Summary"])
    ws1.append(["Project Name", "Planned Cost", "Actual Cost", "Variance ($)", "Variance (%)"])
    
    for cell in ws1[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in financial_summary['budget_vs_actual']:
        ws1.append([
            row['project_name'],
            row['planned_cost'],
            row['actual_cost'],
            row['variance_amount'],
            row['variance_percent']
        ])
    
    # Sheet 2: EVM Analytics
    ws2 = wb.create_sheet("EVM Analytics")
    ws2.append(["Project KPI Breakdown"])
    ws2.append(["Project", "PV", "EV", "AC", "SV ($)", "CV ($)", "SPI", "CPI", "EAC", "ETC"])
    
    for cell in ws2[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in evm_analytics['project_kpis']:
        ws2.append([
            row['project_name'],
            row['pv'],
            row['ev'],
            row['ac'],
            row['sv'],
            row['cv'],
            row['spi'],
            row['cpi'],
            row['eac'],
            row['etc']
        ])
    
    # Sheet 3: Risk & Forecasts
    ws3 = wb.create_sheet("Risk & Forecasts")
    ws3.append(["Delay Forecast"])
    ws3.append(["Metric", "Value (Days)"])
    ws3.append(["P50 (Median)", risk_forecasts['delay_forecast']['p50']])
    ws3.append(["P90 (90th Percentile)", risk_forecasts['delay_forecast']['p90']])
    ws3.append([])
    
    ws3.append(["Top Risk Items"])
    ws3.append(["Item Name", "Project", "Cost Variance ($)", "Schedule Delay (Days)"])
    
    for cell in ws3[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in risk_forecasts['top_risk_items']:
        ws3.append([
            row['item_name'],
            row['project_name'],
            row['cost_variance'],
            row['schedule_delay']
        ])
    
    # Sheet 4: Operational Performance
    ws4 = wb.create_sheet("Operational Performance")
    ws4.append(["Supplier Scorecard"])
    ws4.append(["Supplier Name", "Total Orders", "On-Time Delivery Rate (%)", "Avg Cost Variance (%)"])
    
    for cell in ws4[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row in operational_performance['supplier_scorecard']:
        ws4.append([
            row['supplier_name'],
            row['total_orders'],
            row['on_time_delivery_rate'],
            row['avg_cost_variance_percent']
        ])
    
    # Auto-size columns
    for ws in [ws1, ws2, ws3, ws4]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/filters/projects")
async def get_projects_for_filter(
    current_user: User = Depends(require_analytics_access()),
    db: AsyncSession = Depends(get_db)
):
    """Get list of projects for filter dropdown."""
    result = await db.execute(select(Project).where(Project.is_active == True))
    projects = result.scalars().all()
    
    project_list = [{'id': p.id, 'name': p.name, 'code': p.project_code} for p in projects]
    print(f"📋 DEBUG: Returning {len(project_list)} projects for filter")
    return project_list


@router.get("/filters/suppliers")
async def get_suppliers_for_filter(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_analytics_access())
):
    """Get list of suppliers for filter dropdown."""
    from app.models import ProcurementOption
    
    # Get unique supplier names from procurement options that have finalized decisions
    result = await db.execute(
        select(ProcurementOption.supplier_name)
        .where(ProcurementOption.is_active == True)
        .distinct()
        .order_by(ProcurementOption.supplier_name)
    )
    suppliers = result.scalars().all()
    
    # Return as list with name as both id and name (for filter purposes)
    supplier_list = [{'id': s, 'name': s} for s in suppliers if s]
    print(f"🏢 DEBUG: Returning {len(supplier_list)} suppliers for filter")
    return supplier_list


@router.get("/data-summary")
async def get_data_summary(
    current_user: User = Depends(require_analytics_access()),
    db: AsyncSession = Depends(get_db)
):
    """
    Get comprehensive data summary statistics for the Data Summary tab.
    Shows what data exists and what's needed for better analytics.
    """
    from app.models import ProcurementOption
    
    # Get all locked decisions
    locked_result = await db.execute(
        select(FinalizedDecision).where(FinalizedDecision.status == 'LOCKED')
    )
    all_locked = locked_result.scalars().all()
    total_locked = len(all_locked)
    
    # Count items with various data
    with_invoice = sum(1 for d in all_locked if d.actual_invoice_issue_date)
    with_payment = sum(1 for d in all_locked if d.actual_payment_date)
    with_pm_acceptance = sum(1 for d in all_locked if d.pm_accepted_at)
    with_delivery_complete = sum(1 for d in all_locked if d.delivery_status == 'DELIVERY_COMPLETE')
    
    # Count projects
    projects_result = await db.execute(
        select(Project).where(Project.is_active == True)
    )
    total_projects = len(projects_result.scalars().all())
    
    # Count suppliers
    suppliers_result = await db.execute(
        select(ProcurementOption.supplier_name).distinct()
    )
    total_suppliers = len(suppliers_result.scalars().all())
    
    # Count cashflow events
    cashflow_result = await db.execute(
        select(CashflowEvent).where(
            CashflowEvent.event_type == 'INFLOW',
            CashflowEvent.forecast_type == 'ACTUAL',
            CashflowEvent.is_cancelled == False
        )
    )
    total_inflow_events = len(cashflow_result.scalars().all())
    
    # Calculate percentages
    invoice_percent = (with_invoice / total_locked * 100) if total_locked > 0 else 0
    payment_percent = (with_payment / total_locked * 100) if total_locked > 0 else 0
    pm_acceptance_percent = (with_pm_acceptance / total_locked * 100) if total_locked > 0 else 0
    delivery_complete_percent = (with_delivery_complete / total_locked * 100) if total_locked > 0 else 0
    
    # Calculate data quality score
    data_quality_score = (
        (payment_percent * 0.3) +  # 30% weight on payments
        (pm_acceptance_percent * 0.4) +  # 40% weight on PM acceptance (most important for EVM)
        (invoice_percent * 0.2) +  # 20% weight on invoices
        (delivery_complete_percent * 0.1)  # 10% weight on delivery complete
    )
    
    # Determine status
    if data_quality_score >= 80:
        quality_status = 'excellent'
    elif data_quality_score >= 60:
        quality_status = 'good'
    elif data_quality_score >= 40:
        quality_status = 'fair'
    elif data_quality_score >= 20:
        quality_status = 'limited'
    else:
        quality_status = 'minimal'
    
    return {
        'overall': {
            'total_locked_items': total_locked,
            'total_projects': total_projects,
            'total_suppliers': total_suppliers,
            'data_quality_score': round(data_quality_score, 1),
            'quality_status': quality_status
        },
        'actuals_data': {
            'with_invoice': {
                'count': with_invoice,
                'percent': round(invoice_percent, 1)
            },
            'with_payment': {
                'count': with_payment,
                'percent': round(payment_percent, 1)
            },
            'with_pm_acceptance': {
                'count': with_pm_acceptance,
                'percent': round(pm_acceptance_percent, 1)
            },
            'with_delivery_complete': {
                'count': with_delivery_complete,
                'percent': round(delivery_complete_percent, 1)
            },
            'cashflow_inflow_events': total_inflow_events
        },
        'report_readiness': {
            'financial_summary': 'good' if payment_percent >= 5 else 'limited',
            'evm_analytics': 'good' if pm_acceptance_percent >= 5 else 'limited',
            'risk_forecasts': 'good' if payment_percent >= 5 else 'limited',
            'operational_performance': 'good' if pm_acceptance_percent >= 5 else 'limited'
        },
        'recommendations': [
            {
                'priority': 'high',
                'action': 'Increase PM Acceptances',
                'current': with_pm_acceptance,
                'target': max(20, int(total_locked * 0.1)),
                'impact': 'Unlocks full EVM analytics (EV, CPI, SPI)'
            } if pm_acceptance_percent < 10 else None,
            {
                'priority': 'medium',
                'action': 'Continue Payment Registration',
                'current': with_payment,
                'target': max(50, int(total_locked * 0.2)),
                'impact': 'Improves cash flow and AC metrics'
            } if payment_percent < 20 else None,
            {
                'priority': 'medium',
                'action': 'Complete Delivery Workflow',
                'current': with_delivery_complete,
                'target': with_pm_acceptance,
                'impact': 'Enables invoice entry and cycle time analysis'
            } if with_delivery_complete < with_pm_acceptance else None,
        ]
    }

