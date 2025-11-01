"""
Procurement Plan & Delivery Tracking Router
Handles the operational tracking of finalized procurement decisions
with role-based access control for Procurement Team and Project Managers.
"""
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func
from sqlalchemy.orm import selectinload
from typing import List, Dict, Any, Optional
from datetime import datetime, date
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.models import FinalizedDecision, User, Project, ProjectItem, ProcurementOption, ProjectAssignment, CashflowEvent, SupplierPayment
from app.models_invoice_payment import Invoice, Payment, PaymentStatus
from app.schemas import (
    FinalizedDecision as FinalizedDecisionSchema,
    ProcurementDeliveryConfirmationRequest,
    PMDeliveryAcceptanceRequest,
    ActualInvoiceDataRequest
)
from app.auth import get_current_user, require_role

router = APIRouter(prefix="/procurement-plan", tags=["Procurement Plan"])
logger = logging.getLogger(__name__)


async def _batch_calculate_payment_statuses(decision_ids: List[int], db: AsyncSession) -> Dict[int, Dict[str, str]]:
    """
    Batch calculate Payment In and Payment Out statuses for multiple decisions.
    Returns: {decision_id: {'payment_in_status': '...', 'payment_out_status': '...'}}
    """
    if not decision_ids:
        return {}
    
    # Initialize all statuses to 'not_paid'
    statuses_map = {did: {'payment_in_status': 'not_paid', 'payment_out_status': 'not_paid'} for did in decision_ids}
    
    try:
        # Batch query all invoices for these decisions
        invoices_result = await db.execute(
            select(Invoice).where(Invoice.decision_id.in_(decision_ids))
        )
        invoices = invoices_result.scalars().all()
        
        # Create invoice_id -> decision_id mapping
        invoice_to_decision = {inv.id: inv.decision_id for inv in invoices}
        invoice_ids = [inv.id for inv in invoices]
        
        # Batch query all payments for these invoices
        if invoice_ids:
            payments_result = await db.execute(
                select(Payment).where(
                    Payment.invoice_id.in_(invoice_ids),
                    Payment.status == PaymentStatus.COMPLETED
                )
            )
            payments = payments_result.scalars().all()
            
            # Calculate total payments per invoice
            payments_by_invoice = {}
            for payment in payments:
                if payment.invoice_id not in payments_by_invoice:
                    payments_by_invoice[payment.invoice_id] = 0
                payments_by_invoice[payment.invoice_id] += float(payment.payment_amount)
            
            # Update Payment In statuses
            for invoice in invoices:
                decision_id = invoice.decision_id
                invoice_amount = float(invoice.invoice_amount)
                total_paid = payments_by_invoice.get(invoice.id, 0)
                
                if total_paid >= invoice_amount:
                    statuses_map[decision_id]['payment_in_status'] = 'fully_paid'
                elif total_paid > 0:
                    statuses_map[decision_id]['payment_in_status'] = 'partially_paid'
    except Exception as e:
        # If Invoice/Payment tables don't exist or models aren't accessible, skip Payment In calculation
        logger.warning(f"Error calculating Payment In statuses: {e}")
    
    try:
        # Batch query all supplier payments for these decisions
        supplier_payments_result = await db.execute(
            select(SupplierPayment).where(
                SupplierPayment.decision_id.in_(decision_ids),
                SupplierPayment.status == 'completed'
            )
        )
        supplier_payments = supplier_payments_result.scalars().all()
        
        # Batch query decisions to get final_cost
        decisions_result = await db.execute(
            select(FinalizedDecision).where(FinalizedDecision.id.in_(decision_ids))
        )
        decisions = decisions_result.scalars().all()
        decisions_dict = {d.id: d for d in decisions}
        
        # Calculate total supplier payments per decision
        payments_by_decision = {}
        for sp in supplier_payments:
            if sp.decision_id not in payments_by_decision:
                payments_by_decision[sp.decision_id] = 0
            payments_by_decision[sp.decision_id] += float(sp.payment_amount)
        
        # Update Payment Out statuses
        for decision_id, decision in decisions_dict.items():
            total_paid_out = payments_by_decision.get(decision_id, 0)
            final_cost = float(decision.final_cost_amount) if decision.final_cost_amount else float(decision.final_cost)
            
            if total_paid_out >= final_cost:
                statuses_map[decision_id]['payment_out_status'] = 'fully_paid'
            elif total_paid_out > 0:
                statuses_map[decision_id]['payment_out_status'] = 'partially_paid'
    except Exception as e:
        # If SupplierPayment table doesn't exist or there's an error, skip Payment Out calculation
        logger.warning(f"Error calculating Payment Out statuses: {e}")
    
    return statuses_map


async def _calculate_payment_statuses(decision_id: int, db: AsyncSession) -> Dict[str, str]:
    """
    Calculate Payment In and Payment Out statuses for a decision.
    Returns: {'payment_in_status': '...', 'payment_out_status': '...'}
    """
    payment_in_status = "not_paid"
    payment_out_status = "not_paid"
    
    try:
        # Calculate Payment In (Customer Payments - Revenue)
        invoice_result = await db.execute(
            select(Invoice).where(Invoice.decision_id == decision_id)
        )
        invoice = invoice_result.scalars().first()
        
        if invoice:
            # Calculate total payments received for this invoice
            payments_result = await db.execute(
                select(Payment).where(
                    Payment.invoice_id == invoice.id,
                    Payment.status == PaymentStatus.COMPLETED
                )
            )
            payments = payments_result.scalars().all()
            total_paid_in = sum(float(p.payment_amount) for p in payments)
            invoice_amount = float(invoice.invoice_amount)
            
            if total_paid_in >= invoice_amount:
                payment_in_status = "fully_paid"
            elif total_paid_in > 0:
                payment_in_status = "partially_paid"
            else:
                payment_in_status = "not_paid"
    except Exception as e:
        logger.warning(f"Error calculating Payment In status for decision {decision_id}: {e}")
    
    try:
        # Calculate Payment Out (Supplier Payments - Costs)
        supplier_payments_result = await db.execute(
            select(SupplierPayment).where(
                SupplierPayment.decision_id == decision_id,
                SupplierPayment.status == 'completed'
            )
        )
        supplier_payments = supplier_payments_result.scalars().all()
        
        if supplier_payments:
            # Get the decision to compare with final_cost
            decision_result = await db.execute(
                select(FinalizedDecision).where(FinalizedDecision.id == decision_id)
            )
            decision = decision_result.scalar_one_or_none()
            
            if decision:
                total_paid_out = sum(float(sp.payment_amount) for sp in supplier_payments)
                final_cost = float(decision.final_cost_amount) if decision.final_cost_amount else float(decision.final_cost)
                
                if total_paid_out >= final_cost:
                    payment_out_status = "fully_paid"
                elif total_paid_out > 0:
                    payment_out_status = "partially_paid"
                else:
                    payment_out_status = "not_paid"
    except Exception as e:
        logger.warning(f"Error calculating Payment Out status for decision {decision_id}: {e}")
    
    return {
        'payment_in_status': payment_in_status,
        'payment_out_status': payment_out_status
    }


def _filter_decision_for_role(decision: FinalizedDecision, user_role: str, payment_statuses: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    """
    Filter decision data based on user role.
    
    Procurement Team: Full access to all fields
    PM: Limited access - no financial info, no supplier info
    """
    base_data = {
        "id": decision.id,
        "item_code": decision.item_code,
        "project_id": decision.project_id,
        "quantity": decision.quantity,
        "delivery_date": decision.delivery_date,
        "delivery_status": decision.delivery_status,
        "actual_delivery_date": decision.actual_delivery_date,
        "serial_number": decision.serial_number,
        "customer_delivery_date": decision.customer_delivery_date,
    }
    
    if user_role in ['procurement', 'admin', 'finance']:
        # Full access
        base_data.update({
            "final_cost": float(decision.final_cost_amount) if decision.final_cost_amount else None,
            "final_cost_currency": decision.final_cost_currency or 'IRR',
            "purchase_date": decision.purchase_date,
            "supplier_name": decision.procurement_option.supplier_name if decision.procurement_option else None,
            "procurement_option_id": decision.procurement_option_id,
            "is_correct_item_confirmed": decision.is_correct_item_confirmed,
            "procurement_delivery_notes": decision.procurement_delivery_notes,
            "procurement_confirmed_at": decision.procurement_confirmed_at,
            "procurement_confirmed_by_id": decision.procurement_confirmed_by_id,
            # Invoice data (with currency support)
            "actual_invoice_issue_date": decision.actual_invoice_issue_date,
            "actual_invoice_amount": float(decision.actual_invoice_amount_value) if decision.actual_invoice_amount_value else (float(decision.actual_invoice_amount) if decision.actual_invoice_amount else None),
            "actual_invoice_currency": decision.actual_invoice_amount_currency or decision.final_cost_currency or 'IRR',
            "actual_invoice_received_date": decision.actual_invoice_received_date,
            "invoice_entered_by_id": decision.invoice_entered_by_id,
            "invoice_entered_at": decision.invoice_entered_at,
            # Payment data
            "actual_payment_amount": float(decision.actual_payment_amount_value) if decision.actual_payment_amount_value else (float(decision.actual_payment_amount) if decision.actual_payment_amount else None),
            "actual_payment_currency": decision.actual_payment_amount_currency or decision.final_cost_currency or 'IRR',
            "actual_payment_date": decision.actual_payment_date,
            "payment_entered_by_id": decision.payment_entered_by_id,
            "payment_entered_at": decision.payment_entered_at,
            # Payment In and Payment Out statuses
            "payment_in_status": payment_statuses.get('payment_in_status', 'not_paid') if payment_statuses else 'not_paid',
            "payment_out_status": payment_statuses.get('payment_out_status', 'not_paid') if payment_statuses else 'not_paid',
        })
    
    if user_role in ['pm', 'pmo', 'admin']:
        # PM-specific fields
        base_data.update({
            "is_accepted_by_pm": decision.is_accepted_by_pm,
            "pm_acceptance_notes": decision.pm_acceptance_notes,
            "pm_accepted_at": decision.pm_accepted_at,
            "pm_accepted_by_id": decision.pm_accepted_by_id,
        })
    
    # Add project and item details
    if decision.project:
        base_data["project_name"] = decision.project.name
        base_data["project_code"] = decision.project.project_code
    
    if decision.project_item:
        base_data["item_name"] = decision.project_item.item_name
        base_data["item_description"] = decision.project_item.description
    
    return base_data


@router.get("/")
async def list_procurement_plan(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    status_filter: Optional[str] = None,
    project_id: Optional[int] = None,
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    invoice_status: Optional[str] = None,
    payment_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Get procurement plan items based on user role with pagination and filtering.
    
    - Procurement Team: All finalized items with full details
    - PM: Only items from their assigned projects, limited details
    """
    from datetime import datetime
    from sqlalchemy import and_, or_, func
    
    # Base query for LOCKED decisions only
    stmt = select(FinalizedDecision).where(
        FinalizedDecision.status == 'LOCKED'
    ).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.procurement_option)
    )
    
    # Apply filters
    filters = []
    
    # Filter by delivery status if provided
    if status_filter:
        filters.append(FinalizedDecision.delivery_status == status_filter)
    
    # Filter by project if provided
    if project_id:
        filters.append(FinalizedDecision.project_id == project_id)
    
    # Filter by invoice status
    if invoice_status == 'invoiced':
        filters.append(FinalizedDecision.actual_invoice_issue_date.isnot(None))
    elif invoice_status == 'not_invoiced':
        filters.append(FinalizedDecision.actual_invoice_issue_date.is_(None))
    
    # Filter by payment status
    if payment_status == 'not_paid':
        filters.append(FinalizedDecision.actual_payment_date.is_(None))
    elif payment_status == 'partially_paid':
        filters.append(and_(
            FinalizedDecision.actual_payment_date.isnot(None),
            FinalizedDecision.actual_payment_amount < FinalizedDecision.actual_invoice_amount
        ))
    elif payment_status == 'fully_paid':
        filters.append(and_(
            FinalizedDecision.actual_payment_date.isnot(None),
            FinalizedDecision.actual_payment_amount >= FinalizedDecision.actual_invoice_amount
        ))
    
    # Filter by date range
    if start_date:
        start_dt = datetime.fromisoformat(start_date)
        filters.append(FinalizedDecision.delivery_date >= start_dt.date())
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        filters.append(FinalizedDecision.delivery_date <= end_dt.date())
    
    # Apply all filters
    if filters:
        stmt = stmt.where(and_(*filters))
    
    # Role-based filtering
    if current_user.role == 'pm':
        # PM: Only see items from their assigned projects
        assigned_projects_stmt = select(ProjectAssignment.project_id).where(
            ProjectAssignment.user_id == current_user.id
        )
        assigned_projects_result = await db.execute(assigned_projects_stmt)
        assigned_project_ids = [row[0] for row in assigned_projects_result.fetchall()]
        
        if not assigned_project_ids:
            return {"items": [], "total": 0, "page": page, "limit": limit}
        
        stmt = stmt.where(FinalizedDecision.project_id.in_(assigned_project_ids))
    
    # Get total count for pagination
    count_stmt = select(func.count(FinalizedDecision.id)).select_from(stmt.subquery())
    total_result = await db.execute(count_stmt)
    total_count = total_result.scalar()
    
    # Apply search filter if provided
    if search:
        search_filters = [
            FinalizedDecision.item_code.ilike(f"%{search}%"),
            FinalizedDecision.project.has(Project.name.ilike(f"%{search}%")),
            FinalizedDecision.project_item.has(ProjectItem.item_name.ilike(f"%{search}%")),
            FinalizedDecision.procurement_option.has(ProcurementOption.supplier_name.ilike(f"%{search}%"))
        ]
        stmt = stmt.where(or_(*search_filters))
    
    # Order by delivery date
    stmt = stmt.order_by(FinalizedDecision.delivery_date.asc())
    
    # Apply pagination
    offset = (page - 1) * limit
    stmt = stmt.offset(offset).limit(limit)
    
    result = await db.execute(stmt)
    decisions = result.scalars().all()
    
    # Batch calculate payment statuses for all decisions
    decision_ids = [d.id for d in decisions]
    payment_statuses_map = await _batch_calculate_payment_statuses(decision_ids, db)
    
    # Filter data based on role
    filtered_data = [
        _filter_decision_for_role(decision, current_user.role, payment_statuses_map.get(decision.id))
        for decision in decisions
    ]
    
    return {
        "items": filtered_data,
        "total": total_count,
        "page": page,
        "limit": limit
    }


@router.get("/{decision_id}")
async def get_procurement_plan_item(
    decision_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get detailed view of a single procurement plan item."""
    stmt = select(FinalizedDecision).where(
        FinalizedDecision.id == decision_id
    ).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.procurement_option),
        selectinload(FinalizedDecision.procurement_confirmed_by),
        selectinload(FinalizedDecision.pm_accepted_by)
    )
    
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Check PM access
    if current_user.role == 'pm':
        # Verify PM is assigned to this project
        assignment_check = await db.execute(
            select(ProjectAssignment).where(
                and_(
                    ProjectAssignment.user_id == current_user.id,
                    ProjectAssignment.project_id == decision.project_id
                )
            )
        )
        if not assignment_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have access to this project's items"
            )
    
    # Calculate payment statuses
    payment_statuses = await _calculate_payment_statuses(decision_id, db)
    
    # Return filtered data
    filtered_data = _filter_decision_for_role(decision, current_user.role, payment_statuses)
    
    # Add user names if available
    if decision.procurement_confirmed_by:
        filtered_data["procurement_confirmed_by_name"] = decision.procurement_confirmed_by.username
    if decision.pm_accepted_by:
        filtered_data["pm_accepted_by_name"] = decision.pm_accepted_by.username
    
    return filtered_data


@router.post("/{decision_id}/confirm-delivery")
async def confirm_delivery(
    decision_id: int,
    confirmation: ProcurementDeliveryConfirmationRequest,
    current_user: User = Depends(require_role(["procurement", "admin"])),
    db: AsyncSession = Depends(get_db)
):
    """
    Procurement Team confirms delivery from supplier.
    Updates delivery status to 'CONFIRMED_BY_PROCUREMENT'.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id)
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    if decision.status != 'LOCKED':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only locked decisions can have delivery confirmed"
        )
    
    # Update delivery confirmation fields
    decision.actual_delivery_date = confirmation.actual_delivery_date
    decision.is_correct_item_confirmed = confirmation.is_correct_item
    decision.serial_number = confirmation.serial_number
    decision.procurement_delivery_notes = confirmation.delivery_notes
    decision.procurement_confirmed_at = datetime.utcnow()
    decision.procurement_confirmed_by_id = current_user.id
    decision.delivery_status = 'CONFIRMED_BY_PROCUREMENT'
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Delivery confirmed successfully",
        "decision_id": decision.id,
        "delivery_status": decision.delivery_status
    }


@router.post("/{decision_id}/accept-delivery")
async def accept_delivery(
    decision_id: int,
    acceptance: PMDeliveryAcceptanceRequest,
    current_user: User = Depends(require_role(["pm", "pmo", "admin"])),
    db: AsyncSession = Depends(get_db)
):
    """
    PM accepts delivery for their project.
    If both procurement confirmed and PM accepted, status becomes 'DELIVERY_COMPLETE'.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id)
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    # Check PM access
    if current_user.role == 'pm':
        assignment_check = await db.execute(
            select(ProjectAssignment).where(
                and_(
                    ProjectAssignment.user_id == current_user.id,
                    ProjectAssignment.project_id == decision.project_id
                )
            )
        )
        if not assignment_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have access to this project"
            )
    
    if decision.delivery_status not in ['CONFIRMED_BY_PROCUREMENT', 'AWAITING_DELIVERY']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot accept delivery with current status: {decision.delivery_status}"
        )
    
    # Update PM acceptance fields
    decision.is_accepted_by_pm = acceptance.is_accepted_for_project
    decision.customer_delivery_date = acceptance.customer_delivery_date
    decision.pm_acceptance_notes = acceptance.acceptance_notes
    decision.pm_accepted_at = datetime.utcnow()
    decision.pm_accepted_by_id = current_user.id
    
    # Check if both confirmations are complete
    if decision.is_correct_item_confirmed and decision.is_accepted_by_pm:
        decision.delivery_status = 'DELIVERY_COMPLETE'
    elif decision.is_accepted_by_pm:
        # PM accepted but procurement hasn't confirmed yet
        decision.delivery_status = 'CONFIRMED_BY_PROCUREMENT'  # Keep current or set to this
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Delivery accepted successfully",
        "decision_id": decision.id,
        "delivery_status": decision.delivery_status
    }


@router.post("/{decision_id}/enter-invoice")
async def enter_invoice_data(
    decision_id: int,
    invoice_data: ActualInvoiceDataRequest,
    current_user: User = Depends(require_role(["procurement", "admin", "finance"])),
    db: AsyncSession = Depends(get_db)
):
    """
    Procurement Team (or Finance/Admin) enters actual invoice data.
    This should only be allowed when delivery_status is 'DELIVERY_COMPLETE'.
    Creates an ACTUAL INFLOW cashflow event.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id).options(
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.project)
    )
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    # Check if delivery is complete
    if decision.delivery_status != 'DELIVERY_COMPLETE':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invoice can only be entered when delivery is complete. Current status: {decision.delivery_status}"
        )
    
    # Check if invoice already entered
    if decision.actual_invoice_issue_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invoice data has already been entered for this decision"
        )
    
    # Update invoice fields
    decision.actual_invoice_issue_date = invoice_data.actual_invoice_issue_date
    decision.actual_invoice_amount = invoice_data.actual_invoice_amount
    decision.actual_invoice_received_date = invoice_data.actual_invoice_received_date
    decision.invoice_entered_by_id = current_user.id
    decision.invoice_entered_at = datetime.utcnow()
    
    if invoice_data.notes:
        decision.notes = (decision.notes or "") + f"\n[Invoice Entry] {invoice_data.notes}"
    
    # Create ACTUAL INFLOW cashflow event
    invoice_date = invoice_data.actual_invoice_received_date or invoice_data.actual_invoice_issue_date
    
    cashflow_event = CashflowEvent(
        related_decision_id=decision.id,
        event_type='INFLOW',
        forecast_type='ACTUAL',
        event_date=invoice_date,
        amount=invoice_data.actual_invoice_amount,
        description=f"Actual invoice received for {decision.item_code} (Project: {decision.project.name if decision.project else 'N/A'})"
    )
    db.add(cashflow_event)
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Invoice data entered successfully",
        "decision_id": decision.id,
        "actual_invoice_amount": float(decision.actual_invoice_amount),
        "cashflow_event_created": True
    }


@router.get("/export/excel")
async def export_procurement_plan(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Export procurement plan to Excel.
    Data filtered based on user role.
    """
    # Get filtered data
    items = await list_procurement_plan(current_user=current_user, db=db)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Plan"
    
    # Define headers based on role
    if current_user.role in ['procurement', 'admin', 'finance']:
        headers = [
            "Item Code", "Item Name", "Project", "Supplier", "Quantity",
            "Final Cost", "Purchase Date", "Planned Delivery", "Actual Delivery",
            "Delivery Status", "Serial Number", "Confirmed", "PM Accepted"
        ]
    else:  # PM
        headers = [
            "Item Code", "Item Name", "Project", "Quantity",
            "Planned Delivery", "Actual Delivery", "Delivery Status",
            "Customer Delivery", "PM Accepted"
        ]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_num, item in enumerate(items, 2):
        if current_user.role in ['procurement', 'admin', 'finance']:
            row_data = [
                item.get("item_code"),
                item.get("item_name"),
                item.get("project_name"),
                item.get("supplier_name"),
                item.get("quantity"),
                item.get("final_cost"),
                item.get("purchase_date"),
                item.get("delivery_date"),
                item.get("actual_delivery_date"),
                item.get("delivery_status"),
                item.get("serial_number"),
                "Yes" if item.get("is_correct_item_confirmed") else "No",
                "Yes" if item.get("is_accepted_by_pm") else "No"
            ]
        else:  # PM
            row_data = [
                item.get("item_code"),
                item.get("item_name"),
                item.get("project_name"),
                item.get("quantity"),
                item.get("delivery_date"),
                item.get("actual_delivery_date"),
                item.get("delivery_status"),
                item.get("customer_delivery_date"),
                "Yes" if item.get("is_accepted_by_pm") else "No"
            ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as streaming response
    filename = f"procurement_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

