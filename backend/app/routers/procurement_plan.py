"""
Procurement Plan & Delivery Tracking Router
Handles the operational tracking of finalized procurement decisions
with role-based access control for Procurement Team and Project Managers.
"""
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.orm import selectinload
from typing import List, Dict, Any, Optional
from datetime import datetime, date
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.models import FinalizedDecision, User, Project, ProjectItem, ProcurementOption, ProjectAssignment, CashflowEvent
from app.schemas import (
    FinalizedDecision as FinalizedDecisionSchema,
    ProcurementDeliveryConfirmationRequest,
    PMDeliveryAcceptanceRequest,
    ActualInvoiceDataRequest
)
from app.auth import get_current_user, require_role

router = APIRouter(prefix="/procurement-plan", tags=["Procurement Plan"])


def _filter_decision_for_role(decision: FinalizedDecision, user_role: str) -> Dict[str, Any]:
    """
    Filter decision data based on user role.
    
    Procurement Team: Full access to all fields
    PM: Limited access - no financial info, no supplier info
    """
    base_data = {
        "id": decision.id,
        "item_code": decision.item_code,
        "project_id": decision.project_id,
        "quantity": decision.quantity,
        "delivery_date": decision.delivery_date,
        "delivery_status": decision.delivery_status,
        "actual_delivery_date": decision.actual_delivery_date,
        "serial_number": decision.serial_number,
        "customer_delivery_date": decision.customer_delivery_date,
    }
    
    if user_role in ['procurement', 'admin', 'finance']:
        # Full access
        base_data.update({
            "final_cost": float(decision.final_cost_amount) if decision.final_cost_amount else None,
            "final_cost_currency": decision.final_cost_currency or 'IRR',
            "purchase_date": decision.purchase_date,
            "supplier_name": decision.procurement_option.supplier_name if decision.procurement_option else None,
            "procurement_option_id": decision.procurement_option_id,
            "is_correct_item_confirmed": decision.is_correct_item_confirmed,
            "procurement_delivery_notes": decision.procurement_delivery_notes,
            "procurement_confirmed_at": decision.procurement_confirmed_at,
            "procurement_confirmed_by_id": decision.procurement_confirmed_by_id,
            # Invoice data (with currency support)
            "actual_invoice_issue_date": decision.actual_invoice_issue_date,
            "actual_invoice_amount": float(decision.actual_invoice_amount_value) if decision.actual_invoice_amount_value else (float(decision.actual_invoice_amount) if decision.actual_invoice_amount else None),
            "actual_invoice_currency": decision.actual_invoice_amount_currency or decision.final_cost_currency or 'IRR',
            "actual_invoice_received_date": decision.actual_invoice_received_date,
            "invoice_entered_by_id": decision.invoice_entered_by_id,
            "invoice_entered_at": decision.invoice_entered_at,
            # Payment data
            "actual_payment_amount": float(decision.actual_payment_amount_value) if decision.actual_payment_amount_value else (float(decision.actual_payment_amount) if decision.actual_payment_amount else None),
            "actual_payment_currency": decision.actual_payment_amount_currency or decision.final_cost_currency or 'IRR',
            "actual_payment_date": decision.actual_payment_date,
            "payment_entered_by_id": decision.payment_entered_by_id,
            "payment_entered_at": decision.payment_entered_at,
        })
    
    if user_role in ['pm', 'pmo', 'admin']:
        # PM-specific fields
        base_data.update({
            "is_accepted_by_pm": decision.is_accepted_by_pm,
            "pm_acceptance_notes": decision.pm_acceptance_notes,
            "pm_accepted_at": decision.pm_accepted_at,
            "pm_accepted_by_id": decision.pm_accepted_by_id,
        })
    
    # Add project and item details
    if decision.project:
        base_data["project_name"] = decision.project.name
        base_data["project_code"] = decision.project.project_code
    
    if decision.project_item:
        base_data["item_name"] = decision.project_item.item_name
        base_data["item_description"] = decision.project_item.description
    
    return base_data


@router.get("/")
async def list_procurement_plan(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    status_filter: Optional[str] = None,
    project_id: Optional[int] = None
):
    """
    Get procurement plan items based on user role.
    
    - Procurement Team: All finalized items with full details
    - PM: Only items from their assigned projects, limited details
    """
    # Base query for LOCKED decisions only
    stmt = select(FinalizedDecision).where(
        FinalizedDecision.status == 'LOCKED'
    ).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.procurement_option)
    )
    
    # Filter by delivery status if provided
    if status_filter:
        stmt = stmt.where(FinalizedDecision.delivery_status == status_filter)
    
    # Role-based filtering
    if current_user.role == 'pm':
        # PM: Only see items from their assigned projects
        assigned_projects_stmt = select(ProjectAssignment.project_id).where(
            ProjectAssignment.user_id == current_user.id
        )
        assigned_projects_result = await db.execute(assigned_projects_stmt)
        assigned_project_ids = [row[0] for row in assigned_projects_result.fetchall()]
        
        if not assigned_project_ids:
            return []
        
        stmt = stmt.where(FinalizedDecision.project_id.in_(assigned_project_ids))
    
    # Additional project filter
    if project_id:
        stmt = stmt.where(FinalizedDecision.project_id == project_id)
    
    # Order by delivery date
    stmt = stmt.order_by(FinalizedDecision.delivery_date.asc())
    
    result = await db.execute(stmt)
    decisions = result.scalars().all()
    
    # Filter data based on role
    filtered_data = [
        _filter_decision_for_role(decision, current_user.role)
        for decision in decisions
    ]
    
    return filtered_data


@router.get("/{decision_id}")
async def get_procurement_plan_item(
    decision_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get detailed view of a single procurement plan item."""
    stmt = select(FinalizedDecision).where(
        FinalizedDecision.id == decision_id
    ).options(
        selectinload(FinalizedDecision.project),
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.procurement_option),
        selectinload(FinalizedDecision.procurement_confirmed_by),
        selectinload(FinalizedDecision.pm_accepted_by)
    )
    
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Check PM access
    if current_user.role == 'pm':
        # Verify PM is assigned to this project
        assignment_check = await db.execute(
            select(ProjectAssignment).where(
                and_(
                    ProjectAssignment.user_id == current_user.id,
                    ProjectAssignment.project_id == decision.project_id
                )
            )
        )
        if not assignment_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have access to this project's items"
            )
    
    # Return filtered data
    filtered_data = _filter_decision_for_role(decision, current_user.role)
    
    # Add user names if available
    if decision.procurement_confirmed_by:
        filtered_data["procurement_confirmed_by_name"] = decision.procurement_confirmed_by.username
    if decision.pm_accepted_by:
        filtered_data["pm_accepted_by_name"] = decision.pm_accepted_by.username
    
    return filtered_data


@router.post("/{decision_id}/confirm-delivery")
async def confirm_delivery(
    decision_id: int,
    confirmation: ProcurementDeliveryConfirmationRequest,
    current_user: User = Depends(require_role(["procurement", "admin"])),
    db: AsyncSession = Depends(get_db)
):
    """
    Procurement Team confirms delivery from supplier.
    Updates delivery status to 'CONFIRMED_BY_PROCUREMENT'.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id)
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    if decision.status != 'LOCKED':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only locked decisions can have delivery confirmed"
        )
    
    # Update delivery confirmation fields
    decision.actual_delivery_date = confirmation.actual_delivery_date
    decision.is_correct_item_confirmed = confirmation.is_correct_item
    decision.serial_number = confirmation.serial_number
    decision.procurement_delivery_notes = confirmation.delivery_notes
    decision.procurement_confirmed_at = datetime.utcnow()
    decision.procurement_confirmed_by_id = current_user.id
    decision.delivery_status = 'CONFIRMED_BY_PROCUREMENT'
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Delivery confirmed successfully",
        "decision_id": decision.id,
        "delivery_status": decision.delivery_status
    }


@router.post("/{decision_id}/accept-delivery")
async def accept_delivery(
    decision_id: int,
    acceptance: PMDeliveryAcceptanceRequest,
    current_user: User = Depends(require_role(["pm", "pmo", "admin"])),
    db: AsyncSession = Depends(get_db)
):
    """
    PM accepts delivery for their project.
    If both procurement confirmed and PM accepted, status becomes 'DELIVERY_COMPLETE'.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id)
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    # Check PM access
    if current_user.role == 'pm':
        assignment_check = await db.execute(
            select(ProjectAssignment).where(
                and_(
                    ProjectAssignment.user_id == current_user.id,
                    ProjectAssignment.project_id == decision.project_id
                )
            )
        )
        if not assignment_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have access to this project"
            )
    
    if decision.delivery_status not in ['CONFIRMED_BY_PROCUREMENT', 'AWAITING_DELIVERY']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot accept delivery with current status: {decision.delivery_status}"
        )
    
    # Update PM acceptance fields
    decision.is_accepted_by_pm = acceptance.is_accepted_for_project
    decision.customer_delivery_date = acceptance.customer_delivery_date
    decision.pm_acceptance_notes = acceptance.acceptance_notes
    decision.pm_accepted_at = datetime.utcnow()
    decision.pm_accepted_by_id = current_user.id
    
    # Check if both confirmations are complete
    if decision.is_correct_item_confirmed and decision.is_accepted_by_pm:
        decision.delivery_status = 'DELIVERY_COMPLETE'
    elif decision.is_accepted_by_pm:
        # PM accepted but procurement hasn't confirmed yet
        decision.delivery_status = 'CONFIRMED_BY_PROCUREMENT'  # Keep current or set to this
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Delivery accepted successfully",
        "decision_id": decision.id,
        "delivery_status": decision.delivery_status
    }


@router.post("/{decision_id}/enter-invoice")
async def enter_invoice_data(
    decision_id: int,
    invoice_data: ActualInvoiceDataRequest,
    current_user: User = Depends(require_role(["procurement", "admin", "finance"])),
    db: AsyncSession = Depends(get_db)
):
    """
    Procurement Team (or Finance/Admin) enters actual invoice data.
    This should only be allowed when delivery_status is 'DELIVERY_COMPLETE'.
    Creates an ACTUAL INFLOW cashflow event.
    """
    stmt = select(FinalizedDecision).where(FinalizedDecision.id == decision_id).options(
        selectinload(FinalizedDecision.project_item),
        selectinload(FinalizedDecision.project)
    )
    result = await db.execute(stmt)
    decision = result.scalar_one_or_none()
    
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    
    # Check if delivery is complete
    if decision.delivery_status != 'DELIVERY_COMPLETE':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invoice can only be entered when delivery is complete. Current status: {decision.delivery_status}"
        )
    
    # Check if invoice already entered
    if decision.actual_invoice_issue_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invoice data has already been entered for this decision"
        )
    
    # Update invoice fields
    decision.actual_invoice_issue_date = invoice_data.actual_invoice_issue_date
    decision.actual_invoice_amount = invoice_data.actual_invoice_amount
    decision.actual_invoice_received_date = invoice_data.actual_invoice_received_date
    decision.invoice_entered_by_id = current_user.id
    decision.invoice_entered_at = datetime.utcnow()
    
    if invoice_data.notes:
        decision.notes = (decision.notes or "") + f"\n[Invoice Entry] {invoice_data.notes}"
    
    # Create ACTUAL INFLOW cashflow event
    invoice_date = invoice_data.actual_invoice_received_date or invoice_data.actual_invoice_issue_date
    
    cashflow_event = CashflowEvent(
        related_decision_id=decision.id,
        event_type='INFLOW',
        forecast_type='ACTUAL',
        event_date=invoice_date,
        amount=invoice_data.actual_invoice_amount,
        description=f"Actual invoice received for {decision.item_code} (Project: {decision.project.name if decision.project else 'N/A'})"
    )
    db.add(cashflow_event)
    
    await db.commit()
    await db.refresh(decision)
    
    return {
        "message": "Invoice data entered successfully",
        "decision_id": decision.id,
        "actual_invoice_amount": float(decision.actual_invoice_amount),
        "cashflow_event_created": True
    }


@router.get("/export/excel")
async def export_procurement_plan(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Export procurement plan to Excel.
    Data filtered based on user role.
    """
    # Get filtered data
    items = await list_procurement_plan(current_user=current_user, db=db)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Plan"
    
    # Define headers based on role
    if current_user.role in ['procurement', 'admin', 'finance']:
        headers = [
            "Item Code", "Item Name", "Project", "Supplier", "Quantity",
            "Final Cost", "Purchase Date", "Planned Delivery", "Actual Delivery",
            "Delivery Status", "Serial Number", "Confirmed", "PM Accepted"
        ]
    else:  # PM
        headers = [
            "Item Code", "Item Name", "Project", "Quantity",
            "Planned Delivery", "Actual Delivery", "Delivery Status",
            "Customer Delivery", "PM Accepted"
        ]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_num, item in enumerate(items, 2):
        if current_user.role in ['procurement', 'admin', 'finance']:
            row_data = [
                item.get("item_code"),
                item.get("item_name"),
                item.get("project_name"),
                item.get("supplier_name"),
                item.get("quantity"),
                item.get("final_cost"),
                item.get("purchase_date"),
                item.get("delivery_date"),
                item.get("actual_delivery_date"),
                item.get("delivery_status"),
                item.get("serial_number"),
                "Yes" if item.get("is_correct_item_confirmed") else "No",
                "Yes" if item.get("is_accepted_by_pm") else "No"
            ]
        else:  # PM
            row_data = [
                item.get("item_code"),
                item.get("item_name"),
                item.get("project_name"),
                item.get("quantity"),
                item.get("delivery_date"),
                item.get("actual_delivery_date"),
                item.get("delivery_status"),
                item.get("customer_delivery_date"),
                "Yes" if item.get("is_accepted_by_pm") else "No"
            ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as streaming response
    filename = f"procurement_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

